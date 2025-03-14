from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import sys
import os
# 获取运行目录
if getattr(sys, 'frozen', False):  # 检测是否是 PyInstaller 打包后的环境
    base_path = sys._MEIPASS
else:
    base_path = os.path.dirname(os.path.abspath(__file__))

#file_path = r"Data\.."
file_dir = os.path.join(base_path, "Data")
# 加载 Excel 文件
file_path = file_dir + r"\原文件.xlsx"
wb = load_workbook(file_path)
ws = wb.active  # 获取默认工作表


# 需要查找的文本
target_text = "序号"

# 遍历所有单元格查找“序号”
target_cell = None
for row in ws.iter_rows():
    for cell in row:
        if cell.value == target_text:
            target_cell = cell
            break
    if target_cell:
        break

# 如果找到了“序号”，获取该列的 B1-B10 单元格数据
if target_cell:
    col_letter = target_cell.column_letter  # 获取列字母
    target_range = f"{col_letter}1:{col_letter}{target_cell.row - 1}"
    data = [ws[f"{col_letter}{r}"].value for r in range(1, target_cell.row)if ws[f"{col_letter}{r}"].value is not None]
    print(f"‘{target_text}’ 位置: {target_cell.coordinate}")
    print(f"获取的单元格范围: {target_range}")
    print("单元格数据:", data)
else:
    print("未找到 ‘序号’")

# 关闭 Excel 文件
wb.close()
