import tkinter as tk
from tkinter import ttk, messagebox
from openpyxl.styles import Font
from openpyxl.styles import Alignment,PatternFill
import os
import sys
from tkinter import Toplevel
from tkinter import filedialog
from package.logger import logger_init,log_message
from package.user_data import JsonDataHandler
from package.excel_handler import ExcelHandler
from package.event_handler import EventHandler
from package.config import CONFIG
# 获取运行目录
if getattr(sys, 'frozen', False):  # 检测是否是 PyInstaller 打包后的环境
    base_path = sys._MEIPASS
else:
    base_path = os.path.dirname(os.path.abspath(__file__))

#file_path = r"Data\.."
file_dir = os.path.join(base_path, "Data")


class MainUI(tk.Tk):
    """
        根据样机出样的测试用例来设计UI
    """
    def __init__(self, event_handler,excel_handler,json_data_handler):
        super().__init__()
        self.event_handler = event_handler  # 接收事件处理类的实例
        self.excel_handler = excel_handler
        self.json_handler = json_data_handler
        self.title(CONFIG['main_window_title'])
        self.geometry(CONFIG['main_window_geometry'])
        self.configure_name = self.excel_handler.get_excel_topic_data(CONFIG['test_modules'])
        self.create_widgets()
        self.states = CONFIG['test_states']
        self.current_index = 0   

    def create_widgets(self):
        # 创建一个主容器框架
        container = ttk.Frame(self)
        container.pack(fill="both", expand=True)
        container.config(borderwidth=2, relief="solid")
        # 创建 Canvas 和垂直滚动条
        self.canvas = tk.Canvas(container,  bd=0, highlightthickness=0)
        v_scrollbar = ttk.Scrollbar(container, orient="vertical", command=self.canvas.yview)
        self.canvas.configure(yscrollcommand=v_scrollbar.set)
        v_scrollbar.pack(side="right", fill="y")
        self.canvas.pack(side="left", fill="both", expand=True)

        # 在 Canvas 内创建一个容器，用于放置所有控件
        self.main_frame = ttk.Frame(self.canvas)
        self.canvas.create_window((0, 0), window=self.main_frame, anchor="nw")

        # 当内部容器尺寸改变时更新 Canvas 的滚动区域
        self.main_frame.bind(
            "<Configure>",
            lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        )
        # 绑定鼠标滚轮事件（Windows系统）
        self.canvas.bind_all("<MouseWheel>", self._on_mousewheel)

        load_frame = tk.LabelFrame(self.main_frame, text="设置", width=730, height=130)
        load_frame.pack(side=tk.TOP, padx=5, pady=5)  
        tk.Button(load_frame, text=CONFIG['load_data_button_name'], width=10, command=self.load_data_button).place(x=10,y=10)
        tk.Button(load_frame, text=CONFIG['save_data_button_name'], width=10, command=self.save_data_button).place(x=110,y=10)
        tk.Button(load_frame, text=CONFIG['creat_report_button_name'], width=14, command=self.creat_report_button).place(x=210,y=10)
        tk.Label(load_frame, text=CONFIG['creat_report_entry_name'], anchor="w").place(x=10,y=50)
        self.creat_report_path_entry = tk.Entry(load_frame, width=50) 
        self.creat_report_path_entry.place(x=130,y=50) 
        tk.Button(load_frame, text=CONFIG['open_dir_button_name'], command=lambda :self.open_report_path(self.creat_report_path_entry,1)).place(x=500,y=50) 
        tk.Label(load_frame, text=CONFIG['cover_report_entry_name'], anchor="w").place(x=10,y=80)
        self.cover_report_path_entry = tk.Entry(load_frame, width=50) 
        self.cover_report_path_entry.place(x=130,y=80)
        tk.Button(load_frame, text=CONFIG['open_file_button_name'], command=lambda :self.open_report_path(self.cover_report_path_entry,0)).place(x=500,y=80) 

        basic_frame = tk.LabelFrame(self.main_frame, text=CONFIG['basic_frame_name'], width=650, height=200)
        basic_frame.pack(side=tk.TOP, padx=5, pady=5)  
        self.excel_title_data = CONFIG['excel_title_list']
        frame,self.salesman_combobox = self.configure_basic_info(basic_frame,"业务员" ,self.excel_title_data[0],1,CONFIG['salesman_list'])
        frame,self.drive_type_entry = self.configure_basic_info(basic_frame,"设备型号",self.excel_title_data[1],0,None)
        self.configure_button(frame,"设备型号",CONFIG['usb_comm_button_name'],self.drive_type_entry)
        frame,self.test_time_entry = self.configure_basic_info(basic_frame,"测试时间",self.excel_title_data[2],0,None)
        self.configure_button(frame,"测试时间","当前时间",self.test_time_entry)
        frame,self.test_person_combobox = self.configure_basic_info(basic_frame,"测试人员",self.excel_title_data[3],1,CONFIG['test_person_list'])
        frame,self.machine_quantity_entry = self.configure_basic_info(basic_frame,"样机数量",self.excel_title_data[4],0,"1")
        self.configure_button(frame,"样机数量","↑",self.machine_quantity_entry)
        frame,self.power_source_combobox = self.configure_basic_info(basic_frame,"电源",self.excel_title_data[5],1,CONFIG['power_list'])
        frame,self.version_entry = self.configure_basic_info(basic_frame,"样机/版本确认",self.excel_title_data[6],0,None)
        self.configure_button(frame,"样机/版本确认",CONFIG['usb_comm_button_name'],self.version_entry) 

        self.test_func_list = []
        self.items_list = list(self.excel_handler.excel_test_list.items()) 
        if not self.items_list:
            return 
        
        if not self.configure_name:
            return 
        
        configure_length = len(self.configure_name)
        
        for i in range(configure_length):
            temp = []
            self.configure_test_component(self.configure_name[i],temp)   
            self.test_func_list.append(temp)
    
    def load_item_data(self,item_list):
        for item in item_list:
            key = item[1]
            data = self.json_handler.find_data(key)
            if data:
                if 'radiobutton' in data:
                    item[2].set(data['radiobutton'])
                if 'entry' in data:
                    item[3].insert(0, data['entry'])
                
    def load_data_button(self):
        data = None
        config = {
            "业务员": (self.excel_title_data[0],1,self.salesman_combobox),
            "设备型号": (self.excel_title_data[1],0,self.drive_type_entry),
            "测试时间": (self.excel_title_data[2],0,self.test_time_entry),
            "测试人员": (self.excel_title_data[3],1,self.test_person_combobox),
            "样机数量": (self.excel_title_data[4],0,self.machine_quantity_entry),
            "电源"    : (self.excel_title_data[5],1,self.power_source_combobox),
            "版本确认": (self.excel_title_data[6],0,self.version_entry),
        }
        for i, (key, flag, commp) in config.items():
            data = self.json_handler.find_data(key)
            if key and data:
                widget_class = commp.winfo_class()  # 获取组件的类名
                if widget_class == "Entry": 
                    commp.delete(0, tk.END)
                    commp.insert(0,data)
                elif widget_class == "TCombobox":  # 如果是 Combobox 组件
                    commp.set(data)
        if not self.test_func_list:
            return 
        totl_len = len(self.test_func_list)
        for i in range(totl_len):
            self.load_item_data(self.test_func_list[i])

    def save_item_data(self,item_list):
        for item in item_list:
            key = item[1]
            if key:
                radiobutton = item[2].get()
                entry = item[3].get()
                value = {"radiobutton":radiobutton,"entry":entry}
                self.json_handler.update_data(key,value) 

    def save_data_button(self):
        config = {
            "config1" : (self.excel_title_data[0],self.salesman_combobox.get()),
            "config2" : (self.excel_title_data[1],self.drive_type_entry.get()),
            "config3" : (self.excel_title_data[2],self.test_time_entry.get()),
            "config4" : (self.excel_title_data[3],self.test_person_combobox.get()),
            "config5" : (self.excel_title_data[4],self.machine_quantity_entry.get()),
            "config6" : (self.excel_title_data[5],self.power_source_combobox.get()),
            "config7" : (self.excel_title_data[6],self.version_entry.get()),
        }
        for i, (key, ui_value) in config.items():
            if key:
                self.json_handler.update_data(key,ui_value)
        
        if not self.test_func_list:
            return 
        
        totl_len = len(self.test_func_list)
        for i in range(totl_len):
            self.save_item_data(self.test_func_list[i])
        self.json_handler.write_json()
        messagebox.showinfo("提示", "保存数据成功")
   
    def open_report_path(self,_entry,flag):
        path = None
        if flag == 1:
            path = filedialog.askdirectory()  # 打开目录选择对话框
        else :
            path = filedialog.askopenfilename(filetypes=[("Excel 文件", "*.xlsx")])
        if path:  # 确保用户选择了文件
            _entry.delete(0, tk.END)  # 清空已有内容
            _entry.insert(0, path)  # 插入文件路径    

    def change_cell_value(self,old_cell,_row,_column):
        column_letter = old_cell[0]
        row_number = old_cell[1:]
        new_column_letter = chr(ord(column_letter) + _column)
        new_row_letter = int(row_number) + _row
        new_cell_coordinate = new_column_letter + str(new_row_letter)
        return new_cell_coordinate
               
    def creat_report_button(self):
        if not self.test_func_list:
            return 
        ws = self.excel_handler.get_ws()
        if ws is None:
            print("Error: 保存后重新加载 Excel 失败")
        # 定义 Excel 数据的映射关系 (title_data 索引, row_offset, col_offset, new_value)
        data_map = {
            0: (0, 0, f"样机/版本外发测试(业务人员：{self.salesman_combobox.get()})"),
            1: (0, 2, self.drive_type_entry.get()),
            2: (0, 2, self.test_time_entry.get()),
            3: (0, 2, self.test_person_combobox.get()),
            4: (0, 2, self.machine_quantity_entry.get()),
            5: (0, 2, self.power_source_combobox.get()),
            6: (0, 2, self.version_entry.get()),
        }

        # 统一处理 Excel 写入逻辑
        for index, (row_offset, col_offset, new_value) in data_map.items():
            cell = self.change_cell_value(
                self.excel_handler.get_cell_value(self.excel_title_data[index]), row_offset, col_offset
            )
            ws[cell] = new_value   

        # 设置单元格填充颜色（黄色）
        yello_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        alignment = Alignment(horizontal="center", vertical="center")
        red_font = Font(color="FF0000")  # 红色的RGB代码是"FF0000"
        new_cell_value = None
        length = len(self.test_func_list)
        for i in range(length):
            total_len = len(self.test_func_list[i])
            _list = self.test_func_list[i]
            for j in range(total_len):
                key = _list[j][0]
                radio_value = _list[j][2].get()
                describe = _list[j][3].get() 
                
                new_cell_coordinate = self.change_cell_value(key,0,1)
                if radio_value == "NONE":
                    new_cell_value  = describe
                else :
                    new_cell_value  = f"{radio_value}\r\n{describe}"
                ws[new_cell_coordinate] = new_cell_value  # 填入列标题
                if radio_value == "NA":
                    ws[new_cell_coordinate].fill = yello_fill
                if radio_value == "NG":
                    ws[new_cell_coordinate].font = red_font
                ws[new_cell_coordinate].alignment = alignment
        # 另存为一个新的 Excel 文件
        path = self.creat_report_path_entry.get()
        drive = self.drive_type_entry.get()
        test_time = self.test_time_entry.get()
        salesman = self.salesman_combobox.get()
        self.excel_handler.save_new_excel(path,drive,test_time,salesman)
        messagebox.showinfo("提示", "创建测试报告成功")

    def configure_button(self,frame,frame_name,_text,_commp):
        self.button_func = {
            "设备型号" : self.event_handler.get_driver_product,
            "测试时间": self.event_handler.get_now_time,
            "样机/版本确认":self.event_handler.get_version_inf,
            "样机数量":self.event_handler.change_machine_num
        }
        if frame_name == '样机数量':
            button = tk.Button(frame, text=_text, command=lambda :self.button_func[frame_name]("add",_commp))
            button.pack(side=tk.LEFT, padx=(20,5), pady=5)  
            button = tk.Button(frame, text="↓", command=lambda :self.button_func[frame_name]("dec",_commp))
            button.pack(side=tk.LEFT, padx=(20,5), pady=5)     
        else :
            button = tk.Button(frame, text=_text, command=lambda :self.button_func[frame_name](_commp))
            button.pack(side=tk.LEFT, padx=(20,5), pady=5)  
            
    def configure_basic_info(self, basic_frame,frame_name,cell_name,comp_type,default_value):
        frame = tk.LabelFrame(basic_frame, text=frame_name, width=730, height=50)
        frame.pack(side=tk.TOP, padx=5, pady=5)
        frame.pack_propagate(False)  # 禁止 LabelFrame 自动调整大小 
        tk.Label(frame, text=cell_name, anchor="w").pack(side=tk.LEFT, padx=5, pady=5) 
        if comp_type == 0:
            entry = tk.Entry(frame, width=30)
            entry.pack(side=tk.LEFT, padx=(20,5), pady=5)
            if default_value:
                entry.insert(0,default_value)
            temp = entry
        elif comp_type == 1:
            combobox = ttk.Combobox(frame, width=8, values=default_value)  # 根据实际情况修改串口号列表
            combobox.pack(side=tk.LEFT, padx=(20,5), pady=5)
            temp = combobox
        else :
            pass
        return frame,temp
    def configure_test_component(self,_name,_list):
        #配置测试组件
        start, range_row = self.excel_handler.get_topic_range(_name)

        if range_row != 0:
            frame = tk.LabelFrame(self.main_frame, text=_name, width=730, height=200)
            frame.pack(side=tk.TOP, padx=5, pady=5)  
            frame.pack_propagate(False)  # 禁止 LabelFrame 自动调整大小  
            items_index = 0
            index_flag = False
            for i in range(len(self.items_list)):
                key, value = self.items_list[items_index]
                if start == key:
                    index_flag = True
                    break
                items_index+=1
            if index_flag:
                for i in range(range_row):
                    key, value = self.items_list[items_index]
                    option, entry = self.standard_test_case(frame, value,i) 
                    _list.append((key,value,option,entry))
                    items_index+=1   
    #创建每个测试项的组件
    def standard_test_case(self,frame,label_name,_row):
        # 通用配置
        padx, pady = 5, 1
        grid_config = {"padx": padx, "pady": pady, "sticky": "w"}
        font_style = ("仿宋", 12)
        _column = 0
        # 输入框组件
        entry = tk.Entry(frame, width=20)
        entry.grid(row=_row, column=_column, **grid_config)
        entry.insert(0, label_name)
        _column += 1
        # 颜色映射
        color_map = {
            "NA": "yellow",
            "NG": "red"
        }
        radio_buttons = {}
        # 状态单选按钮
        status_var = tk.StringVar(value="OK")
        default_bg = None
        def update_bg():
            """更新当前行的 Radiobutton 背景颜色"""
            selected_value = status_var.get()
            for value, rb in radio_buttons.items():
                if value == selected_value:
                    # 选中按钮：使用映射中的颜色
                    rb.config(bg=color_map.get(value, default_bg),
                            activebackground=color_map.get(value, default_bg))
                else:
                    # 未选中按钮恢复为默认白色
                    rb.config(bg=default_bg, activebackground=default_bg)

        # 为状态变量添加追踪：当变量改变时自动调用 update_bg
        status_var.trace_add("write", lambda *args: update_bg())
        for text, value in [("NA", "NA"), ("OK", "OK"), ("NG", "NG"), ("NONE", "NONE")]:
            rb = tk.Radiobutton(frame, text=text, variable=status_var, 
                        value=value, font=font_style,  command=update_bg)
            rb.grid(row=_row, column=_column, **grid_config)
            _column += 1
            radio_buttons[value] = rb

        if default_bg == None:
            default_bg =  rb.cget("bg")
        # 描述组件
        tk.Label(frame, text="描述").grid(row=_row, column=_column, **grid_config)
        describe_entry = tk.Entry(frame, width=20)
        describe_entry.grid(row=_row, column=_column+1, **grid_config)
        _column += 2

        # 全局设置按钮（首行显示）
        if _row == 0:
            tk.Button(frame, text="全部设置成NA/OK/NG", 
                    command=lambda: self.set_all_na_ok_ng(frame.cget("text"))
            ).grid(row=_row, column=_column, **grid_config)
        # 按钮配置字典（核心映射关系）
        button_configs = {
            "电源线供电（走纸电流）": ("走纸20行(USB发送)", self.event_handler.feed_paper_test),
            "电源线供电（平均电流）": ("平均电流(USB发送)", lambda: self.select_inches_window(frame)),
            "电源线供电（峰值电流）": ("打印黑块(USB发送)", lambda:self.select_image_window(frame)),
            "切刀": ("切刀测试(USB发送)", self.event_handler.cut_test),
            "USB线": ("USB接口测试", self.event_handler.usb_comm_test),
            "串口线": ("串口接口测试", lambda:self.select_serial_window(frame)),
            "TTL口": ("TTL接口测试", lambda:self.select_serial_window(frame)),
            "网口": ("网口接口测试", lambda:self.select_ethernet_comm(frame)),
            "4G": ("4G测试",self.event_handler._4g_comm_test),
            "钱箱": ("钱箱接口测试", self.event_handler.cashbox_comm_test),
            "轴侦测（开合盖）": ("状态测试",lambda:self.status_check_comm(frame)),
            "黑标": ("黑标测试", lambda:self.black_label_window(frame)),
            "缝标": ("缝标测试", lambda:self.sew_label_window(frame)),
            "堵纸": ("堵纸/拽纸", lambda:self.jam_pull_paper_window(frame)),
            "过温": ("打印黑块", lambda:self.select_image_window(frame)),
            "切刀回到HOME点": ("发送切刀命令", self.event_handler.cut_cmd),
            "驱动打印票据回收": ("吐纸回收设置", lambda:self.spit_recycing_paper_window(frame)),
            "支持指令协议确认": ("打印自测页", self.event_handler.print_selftest),
            "驱动打自检页样张": ("驱动打印设置", lambda:self.driver_print_window(frame)),
            "小票样张": ("票据样张打印", lambda:self.receipt_print_window(frame)),
            "页模式": ("页模式打印", self.event_handler.print_page_mode),
            "打印宽度确认": ("打印宽度设置", lambda:self.width_test_window(frame)),
            "打印速度确认": ("打印速度设置", lambda:self.speed_test_window(frame)),
        }
        # 动态创建专用按钮
        if config := button_configs.get(label_name):
            text, command = config
            tk.Button(frame, text=text, command=command).grid(
                row=_row, column=_column, **grid_config)
            _column += 1
        return status_var,describe_entry
    #创建新窗口
    def creat_new_window(self,parent,_name,_size):
        # 创建模态窗口
        modal = Toplevel(parent)
        modal.title(_name)
        modal.geometry(_size)
        # 获取鼠标位置
        mouse_x = parent.winfo_pointerx()
        mouse_y = parent.winfo_pointery()
        # 禁止调整窗口大小
        modal.resizable(False, False)
        modal.geometry(f"+{mouse_x}+{mouse_y}")  
        return modal
    #锁定新窗口
    def lock_new_window(self,_modal,_parent):
        _modal.transient(_parent)  # 让对话框依附于父窗口
        _modal.grab_set()  # 模态窗口，锁定主窗口
        _parent.wait_window(_modal)  # 等待窗口关闭
    #选择尺寸窗口
    def select_inches_window(self,parent):
        #创建窗口
        modal = self.creat_new_window(parent,"尺寸选择","150x150",)
        #创建组件
        inches_option = tk.StringVar(value="3寸")
        for text in ("2寸", "3寸"):
            tk.Radiobutton(modal, text=text, variable=inches_option, value=text, font=("仿宋", 12)).pack(padx=5, pady=5)
        tk.Button(modal, text="确认", command=lambda: self.event_handler.avg_I_test(inches_option)).pack(padx=5, pady=5)
        #锁定父窗口
        self.lock_new_window(modal,parent)
    #选择图片大小窗口
    def select_image_window(self,parent):
        modal = self.creat_new_window(parent, "图片选择", "150x150")
        # 创建顶部容器
        top_frame = tk.Frame(modal)
        top_frame.pack(fill="x")
        tk.Label(top_frame, text="图片大小", anchor="w").pack(side="left", padx=5, pady=5)
        image_size = ttk.Combobox(top_frame, width=8, values=['384','576','832'])
        image_size.pack(side="left", padx=5, pady=5)
        image_size.set('576')
        # 独立打包按钮
        tk.Button(modal, text="打印", width=12,command=lambda: self.event_handler.black_print(image_size)).pack(fill="x", padx=5, pady=5)
        self.lock_new_window(modal, parent)
   #选择串口参数窗口
    def select_serial_window(self,parent):
        #创建窗口
        modal = self.creat_new_window(parent,"串口通信测试","150x200",)
        configs = [
            ("端口号", self.event_handler.list_serial_com(), 0),
            ("波特率", ['9600','115200'], '115200'),
            ("流量控制", ['NONE','DstDtr','CtsRts','Xon/Xoff'], 'CtsRts')
        ]
        entries = []
        for i, (text, values, default) in enumerate(configs):
            tk.Label(modal, text=text, anchor="w").grid(row=i,column=0,padx=2,pady=2,sticky="w")
            cb = ttk.Combobox(modal, width=8, values=values)
            cb.grid(row=i,column=0,padx=60,pady=2,sticky="w")
            if values: cb.set(values if text=="端口号" else default)
            entries.append(cb)
        cmd = [
            lambda:self.event_handler.serial_comm_test(entries[0].get(),entries[1].get(),entries[2].get()),
            lambda:self.event_handler.set_baud_rate(entries[1].get())
        ]
        tk.Button(modal, text="打印", width=18, command=cmd[0]).grid(row=i+1,column=0,padx=5,pady=2,sticky="w")
        tk.Button(modal, text="设置波特率(USB设置)",width=18,command=cmd[1]).grid(row=i+2,column=0,padx=5,pady=2,sticky="w")
        #锁定父窗口
        self.lock_new_window(modal,parent)

    #选择网口参数窗口
    def select_ethernet_comm(self,parent):
        modal = self.creat_new_window(parent,"网口测试","180x150")
        fields = [
            {"text": "ip地址:",  "default": "192.168.0.87"},{"text": "端口号:",  "default": "9100"}
        ]
        entries = []
        for idx, field in enumerate(fields):
            tk.Label(modal, text=field["text"], anchor="w").grid(row=idx,column=0,padx=2,pady=2)
            entry = tk.Entry(modal, width=12)
            entry.grid(row=idx,column=1,padx=2,pady=2)
            entry.insert(0, field["default"])
            entries.append(entry)
        # 按钮配置数据 
        func = self.event_handler.ethernet_comm_test
        tk.Button(modal, text="打印", width=8, command=lambda:func(entries[0],entries[1],0)).grid(row=idx+1,column=0,padx=2,pady=2)
        tk.Button(modal, text="设置Ip地址", width=8, command=lambda:func(entries[0],entries[1],1)).grid(row=idx+1,column=1,padx=2,pady=2)
        tk.Button(modal, text="开启DHCP", width=8, command=lambda:func(entries[0],entries[1],2)).grid(row=idx+2,column=0,padx=2,pady=2)
        tk.Button(modal, text="关闭DHCP", width=8, command=lambda:func(entries[0],entries[1],3)).grid(row=idx+2,column=1,padx=2,pady=2)
        #锁定父窗口
        self.lock_new_window(modal,parent)
    #状态检测窗口
    def status_check_comm(self,parent):
        # 创建模态窗口
        modal = Toplevel(parent)
        modal.title("状态检测")
        modal.geometry("200x200")
        # 获取鼠标位置
        mouse_x = parent.winfo_pointerx()
        mouse_y = parent.winfo_pointery()
        # 禁止调整窗口大小
        modal.resizable(False, False)
        modal.geometry(f"+{mouse_x}+{mouse_y}")   

        statu_value_label = tk.Label(modal, text="返回状态值:", anchor="w")
        statu_value_label.place(x=10,y=10)
        statu_value_entry = tk.Entry(modal, width=12) 
        statu_value_entry.place(x=80,y=10)
        status_button = [None]*7
        status_button[6] = tk.Button(modal, text="正常", bg="green",width=12,height=2)
        status_button[6].place(x=50,y=40) 
        
        status_button[0] = tk.Button(modal, text="发送10 04 01", width=12)
        status_button[0].place(x=0,y=100) 
        status_button[1] = tk.Button(modal, text="发送10 04 02", width=12)
        status_button[1].place(x=105,y=100) 
        status_button[2] = tk.Button(modal, text="发送10 04 03", width=12)
        status_button[2].place(x=0,y=130) 
        status_button[3] = tk.Button(modal, text="发送10 04 04", width=12)
        status_button[3].place(x=105,y=130) 
        status_button[4] = tk.Button(modal, text="发送10 04 05", width=12)
        status_button[4].place(x=0,y=160) 
        status_button[5] = tk.Button(modal, text="发送全部", width=12)
        status_button[5].place(x=105,y=160) 
        #statu_value_entry.insert(0,"9100")

        #锁定父窗口
        modal.transient(parent)  # 让对话框依附于父窗口
        modal.grab_set()  # 模态窗口，锁定主窗口
        parent.wait_window(modal)  # 等待窗口关闭    
    #黑标测试窗口
    def black_label_window(self,parent):
        # 创建模态窗口
        modal = Toplevel(parent)
        modal.title("黑标测试")
        modal.geometry("250x160")
        # 获取鼠标位置
        mouse_x = parent.winfo_pointerx()
        mouse_y = parent.winfo_pointery()
        # 禁止调整窗口大小
        modal.resizable(False, False)
        modal.geometry(f"+{mouse_x}+{mouse_y}")   
        black_set_button = [None] * 10 
        black_set_button[0] = tk.Button(modal, text="设置黑标", width=8)
        black_set_button[0].grid(row=0,column=0,pady=5)
        black_set_button[1] = tk.Button(modal, text="取消黑标", width=8)
        black_set_button[1].grid(row=0,column=1,pady=5)
        black_distance_label = tk.Label(modal, text="查找黑标后进纸:", anchor="w")
        black_distance_label.grid(row=1,column=0,pady=5)
        black_distance_entry = tk.Entry(modal, width=6) 
        black_distance_entry.grid(row=1,column=1,pady=5) 
        black_set_button[2] = tk.Button(modal, text="设置", width=6)
        black_set_button[2].grid(row=1,column=2,pady=5) 

        black_set_button[3] = tk.Button(modal, text="查找黑标", width=8)
        black_set_button[3].grid(row=2,column=0,pady=5) 
        black_set_button[4] = tk.Button(modal, text="查找黑标后切纸", width=12)
        black_set_button[4].grid(row=2,column=1,pady=5)  
        black_set_button[5] = tk.Button(modal, text="黑标打印测试", width=12)
        black_set_button[5].grid(row=3,column=0,pady=5)  
        #锁定父窗口
        modal.transient(parent)  # 让对话框依附于父窗口
        modal.grab_set()  # 模态窗口，锁定主窗口
        parent.wait_window(modal)  # 等待窗口关闭    
    #缝标测试窗口
    def sew_label_window(self,parent):
        # 创建模态窗口
        modal = Toplevel(parent)
        modal.title("黑标测试")
        modal.geometry("250x160")
        # 获取鼠标位置
        mouse_x = parent.winfo_pointerx()
        mouse_y = parent.winfo_pointery()
        # 禁止调整窗口大小
        modal.resizable(False, False)
        modal.geometry(f"+{mouse_x}+{mouse_y}")   
        sew_set_button = [None] * 10 
        sew_set_button[0] = tk.Button(modal, text="设置缝标", width=8)
        sew_set_button[0].grid(row=0,column=0,pady=5)
        sew_set_button[1] = tk.Button(modal, text="取消缝标", width=8)
        sew_set_button[1].grid(row=0,column=1,pady=5)
        sew_set_button[2] = tk.Button(modal, text="查找缝标", width=8)
        sew_set_button[2].grid(row=1,column=0,pady=5) 
        sew_set_button[3] = tk.Button(modal, text="查找缝标并切纸", width=12)
        sew_set_button[3].grid(row=1,column=1,pady=5) 

        #锁定父窗口
        modal.transient(parent)  # 让对话框依附于父窗口
        modal.grab_set()  # 模态窗口，锁定主窗口
        parent.wait_window(modal)  # 等待窗口关闭    
    #堵纸/拽纸
    def jam_pull_paper_window(self,parent):
        modal = self.creat_new_window(parent,"堵纸/拽纸测试","150x150")
        #锁定父窗口
        self.lock_new_window(modal,parent)
    #吐纸、回收
    def spit_recycing_paper_window(self,parent):
        modal = self.creat_new_window(parent,"吐纸、回收测试","150x150")
        #锁定父窗口
        self.lock_new_window(modal,parent)
    #驱动打印测试
    def driver_print_window(self,parent):
        modal = self.creat_new_window(parent,"驱动打印测试","150x150")
        #锁定父窗口
        self.lock_new_window(modal,parent)
    #驱动打印测试
    def receipt_print_window(self,parent):
        modal = self.creat_new_window(parent,"票据打印测试","150x150")
        #锁定父窗口
        self.lock_new_window(modal,parent)
    #打印宽度测试
    def width_test_window(self,parent):
        modal = self.creat_new_window(parent,"打印宽度测试","150x150")
        #锁定父窗口
        self.lock_new_window(modal,parent)    
    #打印速度测试
    def speed_test_window(self,parent):
        modal = self.creat_new_window(parent,"打印速度测试","150x150")
        #锁定父窗口
        self.lock_new_window(modal,parent)  
    #设置all窗口
    def set_all_na_ok_ng(self,name): 
        index = self.configure_name.index(name)
        current_index = (self.current_index) % len(self.states)
        self.current_index = self.current_index + 1
        total_len = len(self.test_func_list[index])
        _list = self.test_func_list[index]
        for i in range(total_len):
            _list[i][2].set(self.states[current_index])   

    def _on_mousewheel(self, event):
        # Windows 系统下，event.delta 通常为 120 的倍数
        self.canvas.yview_scroll(-1 * (event.delta // 120), "units")

if __name__ == "__main__":
    #创建logger
    logger_init(base_path)
    # 创建事件处理器实例
    event_handler = EventHandler(base_path,file_dir)
    # 创建Excel处理
    excel_handler = ExcelHandler(base_path,file_dir)
    # 加载json数据
    json_data_handler = JsonDataHandler(file_dir)
    # 创建 UI 实例，并将事件处理器传递给 UI
    app = MainUI(event_handler,excel_handler,json_data_handler)
    app.mainloop()
