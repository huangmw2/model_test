import os
from openpyxl import load_workbook
from package.logger import log_message
from tkinter import  messagebox

class ExcelHandler:
    "处理Excel表格事件"
    def __init__(self,base_path,file_dir):
        self.ws = None
        self.wb = None
        self.base_path = base_path
        self.source_file = os.path.join(file_dir, "原文件.xlsx")
        self.excel_test_list = {}
        try:
            self.wb = load_workbook(self.source_file)
            self.ws = self.wb.active  # 获取当前激活的工作表
            self.loading_test_data()
        except Exception as e:
            messagebox.showerror("错误", f"加载 Execel 数据失败：{e}")
    #获取某列的没项数据 
    def get_excel_topic_data(self,cell_name):
        # 遍历表格查找 "测试项目" 的列索引
        if self.ws == None:
            return None
        col_idx = None
        row_idx = None

        for row in self.ws.iter_rows():
            for cell in row:
                if cell.value == cell_name:
                    col_idx = cell.column  # 获取列索引（数字）
                    row_idx = cell.row     # 获取行索引
                    break
            if col_idx:
                break
        if col_idx == None:
            return None
        column_data = [self.ws.cell(row=i, column=col_idx).value for i in range(row_idx + 1, self.ws.max_row + 1) 
                       if self.ws.cell(row=i, column=col_idx).value is not None]
        return column_data

    def loading_test_data(self):
        # 加载测试内容的数据并存储到列表
        if self.ws == None:
            return 
        col_idx = None
        row_idx = None

        for row in self.ws.iter_rows():
            for cell in row:
                if cell.value == "测试内容":
                    col_idx = cell.column  # 获取列索引（数字）
                    row_idx = cell.row     # 获取行索引
                    break
            if col_idx:
                break
        if col_idx == None:
            return 
        # 获取该列 "测试项目" 下面的所有数据，并存储为字典格式
        column_data = {
            self.ws.cell(row=i, column=col_idx).coordinate: self.ws.cell(row=i, column=col_idx).value
            for i in range(row_idx + 1, self.ws.max_row + 1)
            if self.ws.cell(row=i, column=col_idx).value is not None
        }
        self.excel_test_list = column_data
        return 
    def get_topic_range(self, topic_name):
        #获取每个测试项目共有几项
        for row in self.ws.iter_rows():
            for cell in row:
                if cell.value == topic_name:
                    # 遍历所有合并单元格，检查该单元格是否在合并区域内
                    for merged_range in self.ws.merged_cells.ranges:
                        if cell.coordinate in merged_range:
                            merged_rows = merged_range.max_row - merged_range.min_row + 1  # 计算合并的行数
                            new_col_index = merged_range.max_col + 1  # 右侧一列
                            next_column_cell = self.ws.cell(row=merged_range.min_row, column=new_col_index).coordinate

                            return next_column_cell, merged_rows  # 右侧单元格 + 合并行数

                    # 如果该单元格没有合并，则计算右侧一列的新单元格
                    next_column_cell = self.ws.cell(row=cell.row, column=cell.column + 1).coordinate
                    return next_column_cell, 1  # 单个单元格的合并行数视为 1

        return None, 0  # 未找到该值
    def get_cell_value(self,targer_text):    
        # 遍历所有单元格查找“序号”
        target_cell = None
        for row in self.ws.iter_rows():
            for cell in row:
                if cell.value == targer_text:
                    target_cell = cell
                    break
            if target_cell:
                break
        if target_cell == None:
            return None
        return target_cell.coordinate
    
    def get_ws(self):
        return self.ws
    def save_new_excel(self,path,drive=None,time=None,salesman=None):
        #保存EXCEL文件
        excel_path = self.base_path
        if not path:
            excel_path = self.base_path
        else :
            excel_path = path
        file_name = f"/{drive} 样机版本测试报告_{time}({salesman}).xlsx"
        self.dir_file = excel_path + file_name
        self.wb.save(self.dir_file)
        # 重新加载 Excel，避免读取 None
        self.wb.close()  # 关闭当前 Workbook，释放资源
        self.wb = load_workbook(self.source_file, data_only=True)
        self.ws = self.wb.active 